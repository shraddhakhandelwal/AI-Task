"""
Excel report generator using openpyxl.
Produces a formatted spreadsheet with customer data and solar recommendations.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# Colour palette
GREEN_DARK  = "2C7A2C"
BLUE_DARK   = "1A5276"
GREY_LIGHT  = "F0F0F0"
WHITE       = "FFFFFF"
GREY_MED    = "777777"


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(top=thin, bottom=thin, left=thin, right=thin)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _add_section_header(ws, row: int, title: str):
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(bold=True, size=12, color=WHITE)
    cell.fill = _header_fill(BLUE_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="middle", indent=1)
    ws.row_dimensions[row].height = 24


def _add_data_row(ws, row: int, label: str, value, unit: str = ""):
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.fill = _header_fill(GREY_LIGHT)
    label_cell.border = _thin_border()

    ws.merge_cells(f"B{row}:D{row}")
    val_cell = ws.cell(row=row, column=2, value=value)
    val_cell.border = _thin_border()

    if unit:
        ws.merge_cells(f"E{row}:F{row}")
        unit_cell = ws.cell(row=row, column=5, value=unit)
        unit_cell.font = Font(color=GREY_MED, italic=True)

    ws.row_dimensions[row].height = 22


def _fmt_inr(amount: float) -> str:
    """Format a number in Indian currency notation."""
    try:
        n = int(round(amount))
        s = str(n)
        if len(s) <= 3:
            return f"₹ {s}"
        # Indian grouping: last 3 digits, then groups of 2
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]
        return f"₹ {result.lstrip(',')}"
    except Exception:
        return f"₹ {amount}"


def generate_excel(bill_data: dict, solar: dict) -> bytes:
    """
    Generate a formatted Excel workbook and return it as bytes.

    Args:
        bill_data: dict from extract_bill_data()
        solar:     dict from calculate_solar_recommendation()

    Returns:
        Raw bytes of the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load Calculator"

    # Column widths
    widths = [30, 22, 16, 16, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "ENERGYBAE — Solar Load Calculator"
    title.font = Font(bold=True, size=16, color=WHITE)
    title.fill = _header_fill(GREEN_DARK)
    title.alignment = Alignment(horizontal="center", vertical="middle")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = "Electricity Bill Analysis & Solar System Recommendation"
    sub.font = Font(italic=True, size=11, color="555555")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── Section 1: Customer Info ─────────────────────────────────────────────
    _add_section_header(ws, 4, "SECTION 1 — Customer Information")
    _add_data_row(ws, 5,  "Consumer Name",         bill_data.get("consumer_name", "N/A"))
    _add_data_row(ws, 6,  "Consumer Number",        bill_data.get("consumer_number", "N/A"))
    _add_data_row(ws, 7,  "Meter Number",           bill_data.get("meter_number") or "N/A")
    _add_data_row(ws, 8,  "Distribution Company",   bill_data.get("distribution_company") or "N/A")
    _add_data_row(ws, 9,  "Billing Month",          bill_data.get("billing_month", "N/A"))
    _add_data_row(ws, 10, "Tariff Category",        bill_data.get("tariff_category", "N/A"))

    # ── Section 2: Electricity Usage ─────────────────────────────────────────
    _add_section_header(ws, 12, "SECTION 2 — Electricity Usage")
    _add_data_row(ws, 13, "Units Consumed",           bill_data.get("units_consumed", 0),     "kWh / month")
    _add_data_row(ws, 14, "Sanctioned Load",          bill_data.get("sanctioned_load", 0),    "kW")
    _add_data_row(ws, 15, "Total Bill Amount",        _fmt_inr(bill_data.get("total_bill_amount", 0)), "")
    _add_data_row(ws, 16, "Cost per Unit",            f"₹ {solar.get('cost_per_unit', 0):.2f}",        "₹ / kWh")
    _add_data_row(ws, 17, "Average Daily Consumption", f"{solar.get('daily_units', 0):.2f}",          "kWh / day")

    # ── Section 3: Solar Recommendation ──────────────────────────────────────
    _add_section_header(ws, 19, "SECTION 3 — Solar System Recommendation")

    def bold_row(row, label, value, unit=""):
        _add_data_row(ws, row, label, value, unit)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)

    bold_row(20, "Recommended System Size",  solar.get("recommended_system_size_kw", 0), "kWp")
    bold_row(21, "Estimated Monthly Savings",_fmt_inr(solar.get("estimated_monthly_savings", 0)))
    bold_row(22, "Estimated Annual Savings", _fmt_inr(solar.get("estimated_annual_savings", 0)))
    _add_data_row(ws, 23, "Estimated Payback Period", solar.get("payback_period_years", 0), "years")
    _add_data_row(ws, 24, "CO₂ Reduction",           f"{solar.get('co2_reduction_kg_per_year', 0):,}", "kg CO₂ / year")

    # ── Section 4: Financial Summary ─────────────────────────────────────────
    _add_section_header(ws, 26, "SECTION 4 — Financial Summary (25-year projection)")
    _add_data_row(ws, 27, "Estimated System Cost",   _fmt_inr(solar.get("system_cost_inr", 0)),    "(approx. ₹60,000/kWp installed)")
    _add_data_row(ws, 28, "25-Year Savings",          _fmt_inr(solar.get("lifetime_savings_inr", 0)))
    _add_data_row(ws, 29, "Net Benefit (25yr − Cost)",_fmt_inr(solar.get("net_benefit_inr", 0)))

    # ── Footer ────────────────────────────────────────────────────────────────
    ws.merge_cells("A31:F31")
    foot = ws["A31"]
    foot.value = "Generated by Energybae Solar Load Calculator | www.energybae.in | energybae.co@gmail.com"
    foot.font = Font(italic=True, size=9, color="888888")
    foot.alignment = Alignment(horizontal="center")

    ws.merge_cells("A32:F32")
    disc = ws["A32"]
    disc.value = "* Estimates based on 4.5 peak sun hours/day (India avg.) and ₹60,000/kWp installation cost. Actual results may vary."
    disc.font = Font(italic=True, size=8, color="AAAAAA")
    disc.alignment = Alignment(horizontal="center")

    # Write to buffer and return bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
